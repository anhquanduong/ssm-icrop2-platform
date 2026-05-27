import io
import pandas as pd
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)

def export_history_to_csv(sim_history: dict) -> str:
    """
    Flattens the simulation history dictionary collection into a single,
    cohesive unified Pandas DataFrame and returns it as a raw UTF-8 CSV string.
    """
    if not sim_history:
        logger.warning("Empty simulation history. Returning empty CSV string.")
        return ""
    
    # Vertically concatenate all daily dataframes in the history
    combined_df = pd.concat(list(sim_history.values()), ignore_index=True)
    logger.info(f"Flattened simulation history of size {len(combined_df)} records to CSV.")
    return combined_df.to_csv(index=False)

def export_history_to_xlsx(sim_history: dict, 
                           soil_config: dict, 
                           latitude: float, 
                           longitude: float, 
                           crop_name: str) -> bytes:
    """
    Generates a professionally styled multi-sheet Excel workbook in-memory
    containing:
      - Sheet 1 ("Summary Dashboard"): Active soil configuration, coordinates,
        active crop profile name, and high-level summary metadata rows.
      - Sheet 2 ("Simulation Matrix Data"): Continuous daily records formatted
        with corporate navy headers, alternating zebra stripes, cell borders,
        and precise number masking (integers and two-decimal floats).
        
    Returns:
        bytes: The binary stream value of the Excel sheet workbook.
    """
    if not sim_history:
        logger.warning("Empty simulation history. Returning empty bytes.")
        return b""
        
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    
    # Remove openpyxl's default initial worksheet to enforce custom tabs
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # --- STYLING COEFFICIENTS SYSTEM ---
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_side = Side(style='thin', color='D3D3D3')
    gray_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    font_title = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=11, bold=True)
    font_normal = Font(name="Segoe UI", size=11)
    
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")
    
    # --- WORKSHEET 1: SUMMARY DASHBOARD ---
    ws_summary = wb.create_sheet(title="Summary Dashboard")
    ws_summary.views.sheetView[0].showGridLines = True  # Enable gridlines explicitly
    
    # Render Dashboard Header Block
    ws_summary.cell(row=1, column=1, value="BOKU SSM-iCrop Simulation Summary Report").font = font_title
    
    # Check if any scenario in sim_history was executed in Potential ("Classic") mode
    management_status = "Active"
    for label, df in sim_history.items():
        fidelity = df["Model_Fidelity"].iloc[0] if "Model_Fidelity" in df.columns else "Advanced"
        if "Classic" in str(fidelity) or "Potential" in str(fidelity):
            management_status = "Ignored (Potential Baseline)"
            break

    # Render Coordinates and Soil Profile settings metadata rows
    metadata = [
        ("Metadata Parameter", "Configuration Value"),
        ("Active Crop Profile", crop_name),
        ("Simulation Latitude", round(latitude, 4)),
        ("Simulation Longitude", round(longitude, 4)),
        ("Total Root Zone Depth (mm)", soil_config.get("depth_mm", "-")),
        ("Initial Soil Water Content (% Vol)", soil_config.get("initial_water_percent", "-")),
        ("Plant Available Water Capacity (PAWC) (mm/m)", soil_config.get("pawc_mm_m", "-")),
        ("Soil Organic Matter (SOM) (%)", soil_config.get("som_percent", "-")),
        ("Management", management_status),
    ]
    
    meta_start_row = 3
    for idx, (prop, val) in enumerate(metadata):
        r = meta_start_row + idx
        c1 = ws_summary.cell(row=r, column=1, value=prop)
        c2 = ws_summary.cell(row=r, column=2, value=val)
        
        c1.border = gray_border
        c2.border = gray_border
        
        if idx == 0:
            c1.font = font_header
            c1.fill = navy_fill
            c1.alignment = align_left
            c2.font = font_header
            c2.fill = navy_fill
            c2.alignment = align_left
        else:
            c1.font = font_bold
            c1.alignment = align_left
            c2.font = font_normal
            c2.alignment = align_right
            
            # Apply zebra coloring to summary table rows
            if idx % 2 == 1:
                c1.fill = zebra_fill
                c2.fill = zebra_fill
                
    # Render list of simulated scenarios summary block
    next_row = meta_start_row + len(metadata) + 2
    ws_summary.cell(row=next_row, column=1, value="Registered Simulation Scenarios").font = font_bold
    
    scen_header_row = next_row + 1
    c_sh1 = ws_summary.cell(row=scen_header_row, column=1, value="Scenario Label")
    c_sh1.font = font_header
    c_sh1.fill = navy_fill
    c_sh1.border = gray_border
    c_sh2 = ws_summary.cell(row=scen_header_row, column=2, value="Records Count (Days)")
    c_sh2.font = font_header
    c_sh2.fill = navy_fill
    c_sh2.border = gray_border
    
    for s_idx, (scen_label, scen_df) in enumerate(sim_history.items()):
        curr_scen_row = scen_header_row + 1 + s_idx
        sc1 = ws_summary.cell(row=curr_scen_row, column=1, value=scen_label)
        sc2 = ws_summary.cell(row=curr_scen_row, column=2, value=len(scen_df))
        
        sc1.font = font_normal
        sc1.border = gray_border
        sc1.alignment = align_left
        sc2.font = font_normal
        sc2.border = gray_border
        sc2.alignment = align_right
        
        if s_idx % 2 == 1:
            sc1.fill = zebra_fill
            sc2.fill = zebra_fill
            
    # Auto-adjust column sizes on the Summary Dashboard
    for col in ws_summary.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 18)

    # --- WORKSHEET 2: SIMULATION MATRIX DATA ---
    ws_data = wb.create_sheet(title="Simulation Matrix Data")
    ws_data.views.sheetView[0].showGridLines = True  # Enable gridlines explicitly
    ws_data.freeze_panes = "A2"  # Freeze top row headers
    
    # Flatten history dict into a single dataframe
    combined_df = pd.concat(list(sim_history.values()), ignore_index=True)
    
    # Write dataframe rows including headers
    for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=True)):
        ws_data.append(row)
        curr_row_num = r_idx + 1
        
        # Style row cells
        for c_idx in range(1, len(row) + 1):
            cell = ws_data.cell(row=curr_row_num, column=c_idx)
            cell.border = gray_border
            
            if curr_row_num == 1:
                # Corporate navy headers styling
                cell.font = font_header
                cell.fill = navy_fill
                cell.alignment = align_center
            else:
                cell.font = font_normal
                
                # Apply alternating gray/white zebra striping to rows
                if curr_row_num % 2 == 1:
                    cell.fill = zebra_fill
                else:
                    cell.fill = white_fill
                    
                # Apply exact cell alignments and numeric formatting masks based on column names
                col_name = combined_df.columns[c_idx - 1]
                
                if col_name in ["Scenario", "Model_Fidelity", "Management"]:
                    cell.alignment = align_left
                elif col_name == "DOY":
                    cell.alignment = align_center
                    cell.number_format = "#,##0"  # Integer mask
                elif col_name == "BIOMASS":
                    cell.alignment = align_right
                    cell.number_format = "#,##0"  # Integer mask
                elif col_name == "LAI":
                    cell.alignment = align_right
                    cell.number_format = "0.00"   # Two-decimal mask
                elif col_name in ["F_WATER", "F_NUTR"]:
                    cell.alignment = align_right
                    cell.number_format = "0.00"   # Two-decimal stress mask

    # Auto-adjust column sizes on the Data Worksheet
    for col in ws_data.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max(max_len + 4, 15)

    # Save to BytesIO stream and return bytes
    wb.save(buffer)
    logger.info("Professionally formatted XLSX Excel workbook generated successfully in-memory.")
    return buffer.getvalue()
